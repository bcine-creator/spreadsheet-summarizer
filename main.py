import os
import io
import csv
import requests
import openpyxl
from collections import defaultdict
from typing import Dict, Union
from fastapi import FastAPI, Query
from fastapi.responses import JSONResponse
from pydantic import BaseModel
from openai import OpenAI

app = FastAPI()
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

class SheetSummaries(BaseModel):
    summaries: Dict[str, Union[str, Dict]]

def convert_to_xlsx_url(google_url: str) -> str:
    if "/d/" in google_url:
        file_id = google_url.split("/d/")[1].split("/")[0]
        return f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx"
    raise ValueError("Invalid Google Sheets URL")

@app.get("/summarize", response_model=SheetSummaries)
def summarize_spreadsheet(url: str = Query(..., description="Public Google Sheets URL")):
    try:
        download_url = convert_to_xlsx_url(url)
        response = requests.get(download_url)
        response.raise_for_status()

        wb = openpyxl.load_workbook(io.BytesIO(response.content), data_only=True, keep_links=False)
        summaries = {}

        for sheet_name in wb.sheetnames:
            try:
                sheet = wb[sheet_name]
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    rows.append("\t".join([str(cell) if cell is not None else "" for cell in row]))
                sheet_text = "\n".join(rows)

                # Parse CSV-like text
                totals = {
                    "totalRealizedPL": 0.0,
                    "totalOptionsTraded": 0,
                    "tradesBySymbol": defaultdict(int)
                }

                lines = sheet_text.splitlines()
                reader = csv.DictReader(lines)

                for row in reader:
                    symbol = row.get("Symbol")
                    if symbol:
                        totals["tradesBySymbol"][symbol] += 1

                    try:
                        qty = int(float(row.get("Quantity", "0").replace(",", "").strip()))
                        totals["totalOptionsTraded"] += qty
                    except (ValueError, AttributeError):
                        pass

                    try:
                        pl = float(row.get("Realized P/L", "0").replace("$", "").replace(",", "").strip())
                        totals["totalRealizedPL"] += pl
                    except (ValueError, AttributeError):
                        pass

                totals["tradesBySymbol"] = dict(totals["tradesBySymbol"])
                summaries[sheet_name] = totals

            except Exception as e:
                print(f"❌ Skipping sheet '{sheet_name}' due to error: {e}")
                summaries[sheet_name] = f"Error reading this sheet: {str(e)}"

        return {"summaries": summaries}

    except Exception as e:
        print("❌ Error:", str(e))
        return JSONResponse(status_code=500, content={"error": str(e)})
